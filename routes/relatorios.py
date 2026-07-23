"""Central de relatórios do painel administrativo."""

from datetime import date, datetime, timedelta
from io import BytesIO

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from core import *  # noqa: F401,F403


def _periodo_padrao():
    hoje = date.today()
    return hoje.replace(day=1).isoformat(), hoje.isoformat()


def _filtros():
    inicio_padrao, fim_padrao = _periodo_padrao()
    inicio = request.args.get("inicio", inicio_padrao).strip() or inicio_padrao
    fim = request.args.get("fim", fim_padrao).strip() or fim_padrao
    status = request.args.get("status", "todos").strip().lower()
    servico_id = request.args.get("servico_id", "").strip()
    funcionario_id = request.args.get("funcionario_id", "").strip()

    try:
        datetime.strptime(inicio, "%Y-%m-%d")
        datetime.strptime(fim, "%Y-%m-%d")
    except ValueError:
        inicio, fim = inicio_padrao, fim_padrao

    if inicio > fim:
        inicio, fim = fim, inicio

    return {
        "inicio": inicio,
        "fim": fim,
        "status": status,
        "servico_id": int(servico_id) if servico_id.isdigit() else None,
        "funcionario_id": int(funcionario_id) if funcionario_id.isdigit() else None,
    }


def _buscar_dados(empresa_id, filtros):
    conn = get_connection()
    condicoes = ["a.empresa_id = ?", "a.data BETWEEN ? AND ?"]
    parametros = [empresa_id, filtros["inicio"], filtros["fim"]]

    if filtros["status"] != "todos":
        condicoes.append("LOWER(COALESCE(a.status, 'agendado')) = ?")
        parametros.append(filtros["status"])
    if filtros["servico_id"]:
        condicoes.append("a.servico_id = ?")
        parametros.append(filtros["servico_id"])
    if filtros["funcionario_id"]:
        condicoes.append("a.funcionario_id = ?")
        parametros.append(filtros["funcionario_id"])

    where_sql = " AND ".join(condicoes)
    agendamentos = conn.execute(
        f"""
        SELECT a.id, a.data, a.hora, a.cliente_nome, a.cliente_telefone,
               LOWER(COALESCE(a.status, 'agendado')) AS status,
               COALESCE(a.valor_total, 0) AS valor_total,
               COALESCE(
                   (SELECT GROUP_CONCAT(s2.nome, ' + ')
                    FROM agendamento_servicos ags
                    JOIN servicos s2 ON s2.id = ags.servico_id
                    WHERE ags.agendamento_id = a.id),
                   s.nome
               ) AS servico_nome,
               COALESCE(f.nome, 'Sem profissional') AS funcionario_nome
        FROM agendamentos a
        JOIN servicos s ON s.id = a.servico_id
        LEFT JOIN funcionarios f ON f.id = a.funcionario_id
        WHERE {where_sql}
        ORDER BY a.data DESC, a.hora DESC
        """,
        tuple(parametros),
    ).fetchall()

    servicos = conn.execute(
        "SELECT id, nome FROM servicos WHERE empresa_id=? ORDER BY nome",
        (empresa_id,),
    ).fetchall()
    funcionarios = conn.execute(
        "SELECT id, nome FROM funcionarios WHERE empresa_id=? ORDER BY nome",
        (empresa_id,),
    ).fetchall()

    clientes = conn.execute(
        """
        SELECT COUNT(*) AS total,
               SUM(CASE WHEN COALESCE(ativo,1)=1 THEN 1 ELSE 0 END) AS ativos,
               SUM(CASE WHEN criado_em BETWEEN ? AND datetime(?, '+1 day') THEN 1 ELSE 0 END) AS novos
        FROM clientes WHERE empresa_id=?
        """,
        (filtros["inicio"], filtros["fim"], empresa_id),
    ).fetchone()
    conn.close()

    total = len(agendamentos)
    cancelados = sum(1 for item in agendamentos if item["status"] == "cancelado")
    concluidos = sum(1 for item in agendamentos if item["status"] in {"concluido", "finalizado"})
    faturamento = sum(float(item["valor_total"] or 0) for item in agendamentos if item["status"] != "cancelado")
    ticket = faturamento / concluidos if concluidos else 0

    resumo = {
        "agendamentos": total,
        "concluidos": concluidos,
        "cancelados": cancelados,
        "faturamento": faturamento,
        "ticket_medio": ticket,
        "clientes_total": clientes["total"] or 0,
        "clientes_ativos": clientes["ativos"] or 0,
        "clientes_novos": clientes["novos"] or 0,
    }
    return agendamentos, servicos, funcionarios, resumo


@app.route("/admin/relatorios")
@login_required
def admin_relatorios():
    empresa_id = session["empresa_id"]
    filtros = _filtros()
    agendamentos, servicos, funcionarios, resumo = _buscar_dados(empresa_id, filtros)
    return render_template(
        "admin/relatorios.html",
        filtros=filtros,
        agendamentos=agendamentos,
        servicos=servicos,
        funcionarios=funcionarios,
        resumo=resumo,
    )


def _linhas_exportacao(agendamentos):
    cabecalho = ["Data", "Hora", "Cliente", "WhatsApp", "Serviço", "Profissional", "Status", "Valor"]
    linhas = []
    for item in agendamentos:
        linhas.append([
            item["data"], item["hora"], item["cliente_nome"], item["cliente_telefone"],
            item["servico_nome"], item["funcionario_nome"], item["status"].capitalize(),
            float(item["valor_total"] or 0),
        ])
    return cabecalho, linhas


@app.route("/admin/relatorios/excel")
@login_required
def admin_relatorios_excel():
    empresa_id = session["empresa_id"]
    filtros = _filtros()
    agendamentos, _, _, resumo = _buscar_dados(empresa_id, filtros)
    cabecalho, linhas = _linhas_exportacao(agendamentos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Agendamentos"
    ws.append(["Bytech Agenda — Relatório de Agendamentos"])
    ws.append([f"Período: {filtros['inicio']} a {filtros['fim']}"])
    ws.append([])
    ws.append(cabecalho)
    for celula in ws[4]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="2563EB")
    for linha in linhas:
        ws.append(linha)
    for coluna in ws.columns:
        ws.column_dimensions[coluna[0].column_letter].width = min(max(len(str(c.value or "")) for c in coluna) + 2, 38)
    ws.append([])
    ws.append(["Total de agendamentos", resumo["agendamentos"]])
    ws.append(["Faturamento", resumo["faturamento"]])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"relatorio_agendamentos_{filtros['inicio']}_{filtros['fim']}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/relatorios/pdf")
@login_required
def admin_relatorios_pdf():
    empresa_id = session["empresa_id"]
    filtros = _filtros()
    agendamentos, _, _, resumo = _buscar_dados(empresa_id, filtros)
    cabecalho, linhas = _linhas_exportacao(agendamentos)

    arquivo = BytesIO()
    documento = SimpleDocTemplate(arquivo, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph("Bytech Agenda — Relatório de Agendamentos", estilos["Title"]),
        Paragraph(f"Período: {filtros['inicio']} a {filtros['fim']}", estilos["Normal"]),
        Spacer(1, 12),
    ]
    dados = [cabecalho] + [[str(v) if not isinstance(v, float) else f"R$ {v:,.2f}" for v in linha] for linha in linhas]
    tabela = Table(dados, repeatRows=1, colWidths=[58, 42, 90, 75, 115, 90, 62, 65])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.extend([tabela, Spacer(1, 12), Paragraph(f"Agendamentos: {resumo['agendamentos']} | Faturamento: R$ {resumo['faturamento']:,.2f}", estilos["Normal"])])
    documento.build(elementos)
    arquivo.seek(0)
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"relatorio_agendamentos_{filtros['inicio']}_{filtros['fim']}.pdf",
        mimetype="application/pdf",
    )
